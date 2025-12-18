"""
DataLibrary - Biblioteca de integración de datos para SkuldBot
Soporte para bases de datos industriales, archivos y servicios cloud.
"""

from typing import List, Dict, Any, Optional, Iterator, Union
from robot.api.deco import keyword, library
from dataclasses import dataclass
import json


@dataclass
class ExtractionResult:
    """Resultado de una extracción de datos"""
    records: List[Dict[str, Any]]
    columns: List[str]
    record_count: int
    schema: Optional[Dict[str, Any]] = None
    state: Optional[Dict[str, Any]] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "records": self.records,
            "columns": self.columns,
            "recordCount": self.record_count,
            "schema": self.schema,
            "state": self.state,
        }


@dataclass
class LoadResult:
    """Resultado de una carga de datos"""
    inserted_count: int
    updated_count: int
    error_count: int
    errors: List[str]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "insertedCount": self.inserted_count,
            "updatedCount": self.updated_count,
            "errorCount": self.error_count,
            "errors": self.errors,
        }


@library(scope="GLOBAL", auto_keywords=True)
class DataLibrary:
    """
    Biblioteca de integración de datos para SkuldBot.

    Soporta extracción y carga desde/hacia:
    - Bases de datos: SQL Server, Oracle, PostgreSQL, MySQL, DB2, Snowflake
    - Archivos: CSV, Excel
    - Cloud: S3, SFTP
    - SaaS: Salesforce, REST APIs

    Keywords principales:
    - Extract From Database: Extrae datos de BD con soporte para queries
    - Load To Database: Carga datos en BD
    - Extract From CSV: Lee archivo CSV
    - Load To CSV: Escribe archivo CSV
    - Extract From REST API: Extrae de API REST con paginación
    """

    ROBOT_LIBRARY_SCOPE = "GLOBAL"
    ROBOT_LIBRARY_DOC_FORMAT = "ROBOT"

    # Database connection cache
    _connections: Dict[str, Any] = {}

    def __init__(self):
        self._connections = {}

    # =========================================================================
    # DATABASE EXTRACTION
    # =========================================================================

    @keyword("Extract From Database")
    def extract_from_database(
        self,
        db_type: str,
        host: str,
        database: str,
        username: str,
        password: str,
        query: Optional[str] = None,
        table: Optional[str] = None,
        columns: Optional[str] = None,
        filter: Optional[str] = None,
        limit: Optional[int] = None,
        batch_size: int = 10000,
        mode: str = "memory",
        port: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Extrae datos de una base de datos.

        Args:
            db_type: Tipo de BD (sqlserver, oracle, postgres, mysql, db2, snowflake)
            host: Host del servidor
            database: Nombre de la base de datos
            username: Usuario
            password: Contraseña
            query: Query SQL personalizado (opcional)
            table: Nombre de tabla (si no se usa query)
            columns: Columnas a seleccionar, separadas por coma (opcional)
            filter: Condición WHERE (opcional)
            limit: Límite de records (opcional)
            batch_size: Tamaño de lote para batching
            mode: Modo de extracción (memory, batch, stream)
            port: Puerto (opcional, usa default del DB)

        Returns:
            Dict con records, columns, recordCount

        Example:
            | ${result}= | Extract From Database | postgres | localhost | mydb | user | pass | query=SELECT * FROM customers |
            | ${data}= | Set Variable | ${result}[records] |
        """
        # Build connection string based on db_type
        conn = self._get_connection(db_type, host, database, username, password, port)

        # Build query if not provided
        if not query:
            query = self._build_select_query(table, columns, filter, limit)
        elif limit and "LIMIT" not in query.upper() and "TOP" not in query.upper():
            # Add limit if not already in query
            if db_type == "sqlserver":
                query = query.replace("SELECT", f"SELECT TOP {limit}", 1)
            else:
                query = f"{query} LIMIT {limit}"

        # Execute based on mode
        if mode == "memory":
            result = self._extract_memory(conn, query, db_type)
        elif mode == "batch":
            result = self._extract_batch(conn, query, batch_size, db_type)
        else:  # stream
            result = self._extract_memory(conn, query, db_type)  # For now same as memory

        return result.to_dict()

    def _get_connection(
        self,
        db_type: str,
        host: str,
        database: str,
        username: str,
        password: str,
        port: Optional[int] = None,
    ) -> Any:
        """Get or create database connection"""
        cache_key = f"{db_type}:{host}:{database}:{username}"

        if cache_key in self._connections:
            return self._connections[cache_key]

        conn = None

        if db_type == "postgres":
            import psycopg2
            conn = psycopg2.connect(
                host=host,
                database=database,
                user=username,
                password=password,
                port=port or 5432,
            )
        elif db_type == "mysql":
            import pymysql
            conn = pymysql.connect(
                host=host,
                database=database,
                user=username,
                password=password,
                port=port or 3306,
            )
        elif db_type == "sqlserver":
            import pyodbc
            driver = "{ODBC Driver 17 for SQL Server}"
            conn_str = f"DRIVER={driver};SERVER={host},{port or 1433};DATABASE={database};UID={username};PWD={password}"
            conn = pyodbc.connect(conn_str)
        elif db_type == "oracle":
            import cx_Oracle
            dsn = cx_Oracle.makedsn(host, port or 1521, service_name=database)
            conn = cx_Oracle.connect(user=username, password=password, dsn=dsn)
        elif db_type == "db2":
            import ibm_db
            import ibm_db_dbi
            conn_str = f"DATABASE={database};HOSTNAME={host};PORT={port or 50000};PROTOCOL=TCPIP;UID={username};PWD={password}"
            ibm_conn = ibm_db.connect(conn_str, "", "")
            conn = ibm_db_dbi.Connection(ibm_conn)
        elif db_type == "snowflake":
            import snowflake.connector
            conn = snowflake.connector.connect(
                user=username,
                password=password,
                account=host,  # For Snowflake, host is the account identifier
                database=database,
            )
        else:
            raise ValueError(f"Tipo de base de datos no soportado: {db_type}")

        self._connections[cache_key] = conn
        return conn

    def _build_select_query(
        self,
        table: Optional[str],
        columns: Optional[str],
        filter: Optional[str],
        limit: Optional[int],
    ) -> str:
        """Build SELECT query from components"""
        if not table:
            raise ValueError("Se requiere 'table' o 'query'")

        cols = columns if columns else "*"
        query = f"SELECT {cols} FROM {table}"

        if filter:
            query += f" WHERE {filter}"

        if limit:
            query += f" LIMIT {limit}"

        return query

    def _extract_memory(self, conn: Any, query: str, db_type: str) -> ExtractionResult:
        """Extract all records into memory"""
        cursor = conn.cursor()
        cursor.execute(query)

        # Get column names
        columns = [desc[0] for desc in cursor.description]

        # Fetch all records
        rows = cursor.fetchall()
        records = [dict(zip(columns, row)) for row in rows]

        cursor.close()

        return ExtractionResult(
            records=records,
            columns=columns,
            record_count=len(records),
        )

    def _extract_batch(
        self, conn: Any, query: str, batch_size: int, db_type: str
    ) -> ExtractionResult:
        """Extract records in batches (for large datasets)"""
        cursor = conn.cursor()
        cursor.execute(query)

        columns = [desc[0] for desc in cursor.description]
        all_records = []

        while True:
            rows = cursor.fetchmany(batch_size)
            if not rows:
                break
            batch_records = [dict(zip(columns, row)) for row in rows]
            all_records.extend(batch_records)

        cursor.close()

        return ExtractionResult(
            records=all_records,
            columns=columns,
            record_count=len(all_records),
        )

    # =========================================================================
    # DATABASE LOADING
    # =========================================================================

    @keyword("Load To Database")
    def load_to_database(
        self,
        db_type: str,
        host: str,
        database: str,
        username: str,
        password: str,
        table: str,
        records: List[Dict[str, Any]],
        mode: str = "insert",
        batch_size: int = 5000,
        port: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Carga datos en una base de datos.

        Args:
            db_type: Tipo de BD
            host: Host del servidor
            database: Base de datos
            username: Usuario
            password: Contraseña
            table: Tabla destino
            records: Lista de diccionarios a insertar
            mode: Modo de carga (insert, upsert, replace)
            batch_size: Tamaño de lote
            port: Puerto

        Returns:
            Dict con insertedCount, updatedCount, errorCount, errors

        Example:
            | ${result}= | Load To Database | postgres | localhost | mydb | user | pass | customers | ${records} |
        """
        if not records:
            return LoadResult(0, 0, 0, []).to_dict()

        conn = self._get_connection(db_type, host, database, username, password, port)
        cursor = conn.cursor()

        inserted = 0
        errors = []
        columns = list(records[0].keys())

        # Build insert statement
        placeholders = ", ".join(["%s"] * len(columns))
        cols_str = ", ".join(columns)
        insert_sql = f"INSERT INTO {table} ({cols_str}) VALUES ({placeholders})"

        # Insert in batches
        for i in range(0, len(records), batch_size):
            batch = records[i:i + batch_size]
            for record in batch:
                try:
                    values = [record.get(col) for col in columns]
                    cursor.execute(insert_sql, values)
                    inserted += 1
                except Exception as e:
                    errors.append(str(e))

            conn.commit()

        cursor.close()

        return LoadResult(
            inserted_count=inserted,
            updated_count=0,
            error_count=len(errors),
            errors=errors[:10],  # Limit error messages
        ).to_dict()

    # =========================================================================
    # CSV OPERATIONS
    # =========================================================================

    @keyword("Extract From CSV")
    def extract_from_csv(
        self,
        path: str,
        delimiter: str = ",",
        encoding: str = "utf-8",
        header: bool = True,
        columns: Optional[str] = None,
        limit: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Extrae datos de un archivo CSV.

        Args:
            path: Ruta al archivo CSV
            delimiter: Delimitador (default: ,)
            encoding: Codificación del archivo
            header: Si tiene fila de encabezado
            columns: Nombres de columnas personalizados (si header=False)
            limit: Límite de filas

        Returns:
            Dict con records, columns, recordCount

        Example:
            | ${result}= | Extract From CSV | /path/to/file.csv |
        """
        import csv

        records = []
        col_names = []

        with open(path, "r", encoding=encoding) as f:
            if header:
                reader = csv.DictReader(f, delimiter=delimiter)
                for i, row in enumerate(reader):
                    if limit and i >= limit:
                        break
                    records.append(dict(row))
                col_names = reader.fieldnames or []
            else:
                reader = csv.reader(f, delimiter=delimiter)
                if columns:
                    col_names = [c.strip() for c in columns.split(",")]
                for i, row in enumerate(reader):
                    if limit and i >= limit:
                        break
                    if not col_names:
                        col_names = [f"col_{j}" for j in range(len(row))]
                    records.append(dict(zip(col_names, row)))

        return ExtractionResult(
            records=records,
            columns=col_names,
            record_count=len(records),
        ).to_dict()

    @keyword("Load To CSV")
    def load_to_csv(
        self,
        path: str,
        records: List[Dict[str, Any]],
        delimiter: str = ",",
        encoding: str = "utf-8",
        append: bool = False,
    ) -> Dict[str, Any]:
        """
        Guarda datos en un archivo CSV.

        Args:
            path: Ruta al archivo CSV
            records: Lista de diccionarios
            delimiter: Delimitador
            encoding: Codificación
            append: Si agregar al archivo existente

        Returns:
            Dict con insertedCount

        Example:
            | ${result}= | Load To CSV | /path/to/output.csv | ${records} |
        """
        import csv

        if not records:
            return LoadResult(0, 0, 0, []).to_dict()

        mode = "a" if append else "w"
        fieldnames = list(records[0].keys())

        with open(path, mode, encoding=encoding, newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter)
            if not append or f.tell() == 0:
                writer.writeheader()
            writer.writerows(records)

        return LoadResult(
            inserted_count=len(records),
            updated_count=0,
            error_count=0,
            errors=[],
        ).to_dict()

    # =========================================================================
    # EXCEL OPERATIONS (delegates to ExcelLibrary for file ops)
    # =========================================================================

    @keyword("Extract From Excel")
    def extract_from_excel(
        self,
        path: str,
        sheet: Optional[str] = None,
        header: bool = True,
        columns: Optional[str] = None,
        limit: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Extrae datos de un archivo Excel.

        Args:
            path: Ruta al archivo Excel
            sheet: Nombre de la hoja
            header: Si tiene fila de encabezado
            columns: Nombres de columnas personalizados
            limit: Límite de filas

        Returns:
            Dict con records, columns, recordCount

        Example:
            | ${result}= | Extract From Excel | /path/to/file.xlsx | Sheet1 |
        """
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active

        records = []
        col_names = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                if header:
                    col_names = [str(c) if c else f"col_{j}" for j, c in enumerate(row)]
                    continue
                elif columns:
                    col_names = [c.strip() for c in columns.split(",")]
                else:
                    col_names = [f"col_{j}" for j in range(len(row))]

            if limit and len(records) >= limit:
                break

            records.append(dict(zip(col_names, row)))

        wb.close()

        return ExtractionResult(
            records=records,
            columns=col_names,
            record_count=len(records),
        ).to_dict()

    @keyword("Load To Excel")
    def load_to_excel(
        self,
        path: str,
        records: List[Dict[str, Any]],
        sheet: str = "Sheet1",
    ) -> Dict[str, Any]:
        """
        Guarda datos en un archivo Excel.

        Args:
            path: Ruta al archivo Excel
            records: Lista de diccionarios
            sheet: Nombre de la hoja

        Returns:
            Dict con insertedCount

        Example:
            | ${result}= | Load To Excel | /path/to/output.xlsx | ${records} |
        """
        import openpyxl

        if not records:
            return LoadResult(0, 0, 0, []).to_dict()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet

        # Write headers
        fieldnames = list(records[0].keys())
        for col, name in enumerate(fieldnames, start=1):
            ws.cell(row=1, column=col, value=name)

        # Write data
        for row_idx, record in enumerate(records, start=2):
            for col_idx, key in enumerate(fieldnames, start=1):
                ws.cell(row=row_idx, column=col_idx, value=record.get(key))

        wb.save(path)

        return LoadResult(
            inserted_count=len(records),
            updated_count=0,
            error_count=0,
            errors=[],
        ).to_dict()

    # =========================================================================
    # S3 OPERATIONS
    # =========================================================================

    @keyword("Extract From S3")
    def extract_from_s3(
        self,
        bucket: str,
        key: str,
        aws_access_key: str,
        aws_secret_key: str,
        region: str = "us-east-1",
        file_type: str = "csv",
    ) -> Dict[str, Any]:
        """
        Extrae datos de un archivo en S3.

        Args:
            bucket: Nombre del bucket
            key: Path del archivo en S3
            aws_access_key: AWS Access Key ID
            aws_secret_key: AWS Secret Access Key
            region: Región AWS
            file_type: Tipo de archivo (csv, json)

        Returns:
            Dict con records, columns, recordCount

        Example:
            | ${result}= | Extract From S3 | my-bucket | data/file.csv | ${KEY} | ${SECRET} |
        """
        import boto3
        import tempfile
        import os

        s3 = boto3.client(
            "s3",
            aws_access_key_id=aws_access_key,
            aws_secret_access_key=aws_secret_key,
            region_name=region,
        )

        # Download to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_type}") as tmp:
            s3.download_file(bucket, key, tmp.name)
            tmp_path = tmp.name

        try:
            if file_type == "csv":
                result = self.extract_from_csv(tmp_path)
            elif file_type == "json":
                with open(tmp_path, "r") as f:
                    data = json.load(f)
                if isinstance(data, list):
                    records = data
                else:
                    records = [data]
                columns = list(records[0].keys()) if records else []
                result = ExtractionResult(
                    records=records,
                    columns=columns,
                    record_count=len(records),
                ).to_dict()
            else:
                raise ValueError(f"Tipo de archivo no soportado: {file_type}")
        finally:
            os.unlink(tmp_path)

        return result

    @keyword("Load To S3")
    def load_to_s3(
        self,
        bucket: str,
        key: str,
        records: List[Dict[str, Any]],
        aws_access_key: str,
        aws_secret_key: str,
        region: str = "us-east-1",
        file_type: str = "csv",
    ) -> Dict[str, Any]:
        """
        Sube datos a S3.

        Args:
            bucket: Nombre del bucket
            key: Path destino en S3
            records: Lista de diccionarios
            aws_access_key: AWS Access Key ID
            aws_secret_key: AWS Secret Access Key
            region: Región AWS
            file_type: Tipo de archivo (csv, json)

        Returns:
            Dict con insertedCount

        Example:
            | ${result}= | Load To S3 | my-bucket | output/data.csv | ${records} | ${KEY} | ${SECRET} |
        """
        import boto3
        import tempfile
        import os

        s3 = boto3.client(
            "s3",
            aws_access_key_id=aws_access_key,
            aws_secret_access_key=aws_secret_key,
            region_name=region,
        )

        # Write to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_type}") as tmp:
            tmp_path = tmp.name

        try:
            if file_type == "csv":
                self.load_to_csv(tmp_path, records)
            elif file_type == "json":
                with open(tmp_path, "w") as f:
                    json.dump(records, f)
            else:
                raise ValueError(f"Tipo de archivo no soportado: {file_type}")

            s3.upload_file(tmp_path, bucket, key)
        finally:
            os.unlink(tmp_path)

        return LoadResult(
            inserted_count=len(records),
            updated_count=0,
            error_count=0,
            errors=[],
        ).to_dict()

    # =========================================================================
    # SFTP OPERATIONS
    # =========================================================================

    @keyword("Extract From SFTP")
    def extract_from_sftp(
        self,
        host: str,
        path: str,
        username: str,
        password: Optional[str] = None,
        private_key: Optional[str] = None,
        port: int = 22,
        file_type: str = "csv",
    ) -> Dict[str, Any]:
        """
        Extrae datos de un archivo en SFTP.

        Args:
            host: Host SFTP
            path: Path del archivo
            username: Usuario
            password: Contraseña (o usar private_key)
            private_key: Path a llave privada
            port: Puerto SFTP
            file_type: Tipo de archivo (csv, json)

        Returns:
            Dict con records, columns, recordCount

        Example:
            | ${result}= | Extract From SFTP | sftp.example.com | /data/file.csv | user | pass |
        """
        import paramiko
        import tempfile
        import os

        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        if private_key:
            pkey = paramiko.RSAKey.from_private_key_file(private_key)
            ssh.connect(host, port=port, username=username, pkey=pkey)
        else:
            ssh.connect(host, port=port, username=username, password=password)

        sftp = ssh.open_sftp()

        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_type}") as tmp:
            sftp.get(path, tmp.name)
            tmp_path = tmp.name

        sftp.close()
        ssh.close()

        try:
            if file_type == "csv":
                result = self.extract_from_csv(tmp_path)
            elif file_type == "json":
                with open(tmp_path, "r") as f:
                    data = json.load(f)
                records = data if isinstance(data, list) else [data]
                columns = list(records[0].keys()) if records else []
                result = ExtractionResult(
                    records=records,
                    columns=columns,
                    record_count=len(records),
                ).to_dict()
            else:
                raise ValueError(f"Tipo de archivo no soportado: {file_type}")
        finally:
            os.unlink(tmp_path)

        return result

    @keyword("Load To SFTP")
    def load_to_sftp(
        self,
        host: str,
        path: str,
        records: List[Dict[str, Any]],
        username: str,
        password: Optional[str] = None,
        private_key: Optional[str] = None,
        port: int = 22,
        file_type: str = "csv",
    ) -> Dict[str, Any]:
        """
        Sube datos a SFTP.

        Args:
            host: Host SFTP
            path: Path destino
            records: Lista de diccionarios
            username: Usuario
            password: Contraseña
            private_key: Path a llave privada
            port: Puerto
            file_type: Tipo de archivo

        Returns:
            Dict con insertedCount

        Example:
            | ${result}= | Load To SFTP | sftp.example.com | /output/data.csv | ${records} | user | pass |
        """
        import paramiko
        import tempfile
        import os

        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        if private_key:
            pkey = paramiko.RSAKey.from_private_key_file(private_key)
            ssh.connect(host, port=port, username=username, pkey=pkey)
        else:
            ssh.connect(host, port=port, username=username, password=password)

        sftp = ssh.open_sftp()

        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_type}") as tmp:
            tmp_path = tmp.name

        try:
            if file_type == "csv":
                self.load_to_csv(tmp_path, records)
            elif file_type == "json":
                with open(tmp_path, "w") as f:
                    json.dump(records, f)
            else:
                raise ValueError(f"Tipo de archivo no soportado: {file_type}")

            sftp.put(tmp_path, path)
        finally:
            os.unlink(tmp_path)

        sftp.close()
        ssh.close()

        return LoadResult(
            inserted_count=len(records),
            updated_count=0,
            error_count=0,
            errors=[],
        ).to_dict()

    # =========================================================================
    # REST API OPERATIONS
    # =========================================================================

    @keyword("Extract From REST API")
    def extract_from_rest_api(
        self,
        url: str,
        method: str = "GET",
        headers: Optional[str] = None,
        body: Optional[str] = None,
        auth_type: Optional[str] = None,
        auth_value: Optional[str] = None,
        pagination_type: Optional[str] = None,
        pagination_param: Optional[str] = None,
        data_path: Optional[str] = None,
        limit: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Extrae datos de una REST API con soporte para paginación.

        Args:
            url: URL del endpoint
            method: Método HTTP (GET, POST)
            headers: Headers como JSON string
            body: Body como JSON string (para POST)
            auth_type: Tipo de auth (bearer, basic, api_key)
            auth_value: Valor de auth (token, user:pass, key)
            pagination_type: Tipo de paginación (offset, cursor, link)
            pagination_param: Parámetro de paginación
            data_path: JSONPath al array de datos (ej: "$.data", "$.results")
            limit: Límite de records

        Returns:
            Dict con records, columns, recordCount

        Example:
            | ${result}= | Extract From REST API | https://api.example.com/users | headers={"Accept": "application/json"} |
        """
        import requests

        req_headers = json.loads(headers) if headers else {}

        # Add authentication
        if auth_type == "bearer":
            req_headers["Authorization"] = f"Bearer {auth_value}"
        elif auth_type == "api_key":
            req_headers["X-API-Key"] = auth_value

        all_records = []
        page = 1
        offset = 0
        next_cursor = None

        while True:
            # Build URL with pagination
            paginated_url = url
            if pagination_type == "offset":
                sep = "&" if "?" in url else "?"
                paginated_url = f"{url}{sep}{pagination_param}={offset}"
            elif pagination_type == "page":
                sep = "&" if "?" in url else "?"
                paginated_url = f"{url}{sep}{pagination_param}={page}"
            elif pagination_type == "cursor" and next_cursor:
                sep = "&" if "?" in url else "?"
                paginated_url = f"{url}{sep}{pagination_param}={next_cursor}"

            # Make request
            if method.upper() == "GET":
                response = requests.get(paginated_url, headers=req_headers)
            else:
                req_body = json.loads(body) if body else None
                response = requests.post(paginated_url, headers=req_headers, json=req_body)

            response.raise_for_status()
            data = response.json()

            # Extract records from response
            if data_path:
                # Simple JSONPath extraction
                path_parts = data_path.replace("$.", "").split(".")
                records = data
                for part in path_parts:
                    if part and records:
                        records = records.get(part, [])
            else:
                records = data if isinstance(data, list) else [data]

            if not records:
                break

            all_records.extend(records)

            # Check limit
            if limit and len(all_records) >= limit:
                all_records = all_records[:limit]
                break

            # Handle pagination
            if pagination_type == "offset":
                offset += len(records)
                if len(records) < 100:  # Assume page size of 100
                    break
            elif pagination_type == "page":
                page += 1
                if len(records) < 100:
                    break
            elif pagination_type == "cursor":
                next_cursor = data.get("next_cursor") or data.get("cursor")
                if not next_cursor:
                    break
            else:
                break  # No pagination

        columns = list(all_records[0].keys()) if all_records else []

        return ExtractionResult(
            records=all_records,
            columns=columns,
            record_count=len(all_records),
        ).to_dict()

    # =========================================================================
    # SALESFORCE OPERATIONS
    # =========================================================================

    @keyword("Extract From Salesforce")
    def extract_from_salesforce(
        self,
        username: str,
        password: str,
        security_token: str,
        query: str,
        domain: str = "login",
    ) -> Dict[str, Any]:
        """
        Extrae datos de Salesforce usando SOQL.

        Args:
            username: Usuario Salesforce
            password: Contraseña
            security_token: Security Token
            query: Query SOQL
            domain: Dominio (login, test)

        Returns:
            Dict con records, columns, recordCount

        Example:
            | ${result}= | Extract From Salesforce | user@company.com | pass | token | SELECT Id, Name FROM Account |
        """
        from simple_salesforce import Salesforce

        sf = Salesforce(
            username=username,
            password=password,
            security_token=security_token,
            domain=domain,
        )

        result = sf.query_all(query)
        records = result.get("records", [])

        # Remove Salesforce metadata from records
        clean_records = []
        for record in records:
            clean = {k: v for k, v in record.items() if not k.startswith("attributes")}
            clean_records.append(clean)

        columns = list(clean_records[0].keys()) if clean_records else []

        return ExtractionResult(
            records=clean_records,
            columns=columns,
            record_count=len(clean_records),
        ).to_dict()

    # =========================================================================
    # BIGQUERY OPERATIONS (Target only)
    # =========================================================================

    @keyword("Load To BigQuery")
    def load_to_bigquery(
        self,
        project: str,
        dataset: str,
        table: str,
        records: List[Dict[str, Any]],
        credentials_json: str,
        mode: str = "append",
    ) -> Dict[str, Any]:
        """
        Carga datos en Google BigQuery.

        Args:
            project: ID del proyecto GCP
            dataset: Dataset de BigQuery
            table: Tabla destino
            records: Lista de diccionarios
            credentials_json: Path al archivo de credenciales JSON
            mode: Modo de escritura (append, truncate)

        Returns:
            Dict con insertedCount

        Example:
            | ${result}= | Load To BigQuery | my-project | my_dataset | my_table | ${records} | /path/to/creds.json |
        """
        from google.cloud import bigquery
        from google.oauth2 import service_account

        credentials = service_account.Credentials.from_service_account_file(
            credentials_json
        )

        client = bigquery.Client(credentials=credentials, project=project)

        table_id = f"{project}.{dataset}.{table}"

        write_disposition = (
            bigquery.WriteDisposition.WRITE_APPEND
            if mode == "append"
            else bigquery.WriteDisposition.WRITE_TRUNCATE
        )

        job_config = bigquery.LoadJobConfig(
            write_disposition=write_disposition,
            source_format=bigquery.SourceFormat.NEWLINE_DELIMITED_JSON,
        )

        job = client.load_table_from_json(records, table_id, job_config=job_config)
        job.result()  # Wait for completion

        return LoadResult(
            inserted_count=len(records),
            updated_count=0,
            error_count=0,
            errors=[],
        ).to_dict()

    # =========================================================================
    # CLEANUP
    # =========================================================================

    @keyword("Close All Data Connections")
    def close_all_connections(self):
        """
        Cierra todas las conexiones de base de datos abiertas.

        Example:
            | Close All Data Connections |
        """
        for key, conn in self._connections.items():
            try:
                conn.close()
            except Exception:
                pass
        self._connections.clear()

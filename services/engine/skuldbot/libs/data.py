"""
DataLibrary - Biblioteca de integración de datos para SkuldBot
Soporte para bases de datos industriales, archivos y servicios cloud.
"""

from typing import List, Dict, Any, Optional, Iterator, Union
from robot.api.deco import keyword, library
from dataclasses import dataclass
from pathlib import Path
from datetime import datetime, date, timezone
from decimal import Decimal, InvalidOperation
import hashlib
import json


@dataclass
class ExtractionResult:
    """Resultado de una extracción de datos"""
    records: List[Dict[str, Any]]
    columns: List[str]
    record_count: int
    schema: Optional[Dict[str, Any]] = None
    state: Optional[Dict[str, Any]] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "records": self.records,
            "columns": self.columns,
            "recordCount": self.record_count,
            "schema": self.schema,
            "state": self.state,
        }


@dataclass
class LoadResult:
    """Resultado de una carga de datos"""
    inserted_count: int
    updated_count: int
    error_count: int
    errors: List[str]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "insertedCount": self.inserted_count,
            "updatedCount": self.updated_count,
            "errorCount": self.error_count,
            "errors": self.errors,
        }


@library(scope="GLOBAL", auto_keywords=True)
class DataLibrary:
    """
    Biblioteca de integración de datos para SkuldBot.

    Soporta extracción y carga desde/hacia:
    - Bases de datos: SQL Server, Oracle, PostgreSQL, MySQL, DB2, Snowflake
    - Archivos: CSV, Excel
    - Cloud: S3, SFTP
    - SaaS: Salesforce, REST APIs

    Keywords principales:
    - Extract From Database: Extrae datos de BD con soporte para queries
    - Load To Database: Carga datos en BD
    - Extract From CSV: Lee archivo CSV
    - Load To CSV: Escribe archivo CSV
    - Extract From REST API: Extrae de API REST con paginación
    """

    ROBOT_LIBRARY_SCOPE = "GLOBAL"
    ROBOT_LIBRARY_DOC_FORMAT = "ROBOT"

    # Database connection cache
    _connections: Dict[str, Any] = {}

    def __init__(self):
        self._connections = {}

    # =========================================================================
    # DATABASE EXTRACTION
    # =========================================================================

    @keyword("Extract From Database")
    def extract_from_database(
        self,
        db_type: str,
        host: str,
        database: str,
        username: str,
        password: str,
        query: Optional[str] = None,
        table: Optional[str] = None,
        columns: Optional[str] = None,
        filter: Optional[str] = None,
        limit: Optional[int] = None,
        batch_size: int = 10000,
        mode: str = "memory",
        port: Optional[int] = None,
        schema: Optional[str] = None,
        warehouse: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Extrae datos de una base de datos.

        Args:
            db_type: Tipo de BD (sqlserver, oracle, postgres, mysql, db2, snowflake)
            host: Host del servidor
            database: Nombre de la base de datos
            username: Usuario
            password: Contraseña
            query: Query SQL personalizado (opcional)
            table: Nombre de tabla (si no se usa query)
            columns: Columnas a seleccionar, separadas por coma (opcional)
            filter: Condición WHERE (opcional)
            limit: Límite de records (opcional)
            batch_size: Tamaño de lote para batching
            mode: Modo de extracción (memory, batch, stream)
            port: Puerto (opcional, usa default del DB)

        Returns:
            Dict con records, columns, recordCount

        Example:
            | ${result}= | Extract From Database | postgres | localhost | mydb | user | pass | query=SELECT * FROM customers |
            | ${data}= | Set Variable | ${result}[records] |
        """
        # Build connection string based on db_type
        conn = self._get_connection(db_type, host, database, username, password, port, schema, warehouse)

        # Build query if not provided
        if not query:
            query = self._build_select_query(table, columns, filter, limit)
        elif limit and "LIMIT" not in query.upper() and "TOP" not in query.upper():
            # Add limit if not already in query
            if db_type == "sqlserver":
                query = query.replace("SELECT", f"SELECT TOP {limit}", 1)
            else:
                query = f"{query} LIMIT {limit}"

        # Execute based on mode
        if mode == "memory":
            result = self._extract_memory(conn, query, db_type)
        elif mode == "batch":
            result = self._extract_batch(conn, query, batch_size, db_type)
        else:  # stream
            result = self._extract_memory(conn, query, db_type)  # For now same as memory

        return result.to_dict()

    def _get_connection(
        self,
        db_type: str,
        host: str,
        database: str,
        username: str,
        password: str,
        port: Optional[int] = None,
        schema: Optional[str] = None,
        warehouse: Optional[str] = None,
    ) -> Any:
        """Get or create database connection"""
        cache_key = f"{db_type}:{host}:{database}:{username}:{schema or ''}:{warehouse or ''}"

        if cache_key in self._connections:
            return self._connections[cache_key]

        conn = None

        if db_type == "postgres":
            import psycopg2
            conn = psycopg2.connect(
                host=host,
                database=database,
                user=username,
                password=password,
                port=port or 5432,
            )
        elif db_type == "mysql":
            import pymysql
            conn = pymysql.connect(
                host=host,
                database=database,
                user=username,
                password=password,
                port=port or 3306,
            )
        elif db_type == "sqlserver":
            import pyodbc
            driver = "{ODBC Driver 17 for SQL Server}"
            conn_str = f"DRIVER={driver};SERVER={host},{port or 1433};DATABASE={database};UID={username};PWD={password}"
            conn = pyodbc.connect(conn_str)
        elif db_type == "oracle":
            import cx_Oracle
            dsn = cx_Oracle.makedsn(host, port or 1521, service_name=database)
            conn = cx_Oracle.connect(user=username, password=password, dsn=dsn)
        elif db_type == "db2":
            import ibm_db
            import ibm_db_dbi
            conn_str = f"DATABASE={database};HOSTNAME={host};PORT={port or 50000};PROTOCOL=TCPIP;UID={username};PWD={password}"
            ibm_conn = ibm_db.connect(conn_str, "", "")
            conn = ibm_db_dbi.Connection(ibm_conn)
        elif db_type == "snowflake":
            import snowflake.connector
            conn = snowflake.connector.connect(
                user=username,
                password=password,
                account=host,  # For Snowflake, host is the account identifier
                database=database,
                schema=schema,
                warehouse=warehouse,
            )
        else:
            raise ValueError(f"Tipo de base de datos no soportado: {db_type}")

        self._connections[cache_key] = conn
        return conn

    def _build_select_query(
        self,
        table: Optional[str],
        columns: Optional[str],
        filter: Optional[str],
        limit: Optional[int],
    ) -> str:
        """Build SELECT query from components"""
        if not table:
            raise ValueError("Se requiere 'table' o 'query'")

        cols = columns if columns else "*"
        query = f"SELECT {cols} FROM {table}"

        if filter:
            query += f" WHERE {filter}"

        if limit:
            query += f" LIMIT {limit}"

        return query

    def _extract_memory(self, conn: Any, query: str, db_type: str) -> ExtractionResult:
        """Extract all records into memory"""
        cursor = conn.cursor()
        cursor.execute(query)

        # Get column names
        columns = [desc[0] for desc in cursor.description]

        # Fetch all records
        rows = cursor.fetchall()
        records = [dict(zip(columns, row)) for row in rows]

        cursor.close()

        return ExtractionResult(
            records=records,
            columns=columns,
            record_count=len(records),
        )

    def _extract_batch(
        self, conn: Any, query: str, batch_size: int, db_type: str
    ) -> ExtractionResult:
        """Extract records in batches (for large datasets)"""
        cursor = conn.cursor()
        cursor.execute(query)

        columns = [desc[0] for desc in cursor.description]
        all_records = []

        while True:
            rows = cursor.fetchmany(batch_size)
            if not rows:
                break
            batch_records = [dict(zip(columns, row)) for row in rows]
            all_records.extend(batch_records)

        cursor.close()

        return ExtractionResult(
            records=all_records,
            columns=columns,
            record_count=len(all_records),
        )

    # =========================================================================
    # DATABASE LOADING
    # =========================================================================

    @keyword("Load To Database")
    def load_to_database(
        self,
        db_type: str,
        host: str,
        database: str,
        username: str,
        password: str,
        table: str,
        records: List[Dict[str, Any]],
        mode: str = "insert",
        batch_size: int = 5000,
        port: Optional[int] = None,
        schema: Optional[str] = None,
        warehouse: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Carga datos en una base de datos.

        Args:
            db_type: Tipo de BD
            host: Host del servidor
            database: Base de datos
            username: Usuario
            password: Contraseña
            table: Tabla destino
            records: Lista de diccionarios a insertar
            mode: Modo de carga (insert, upsert, replace)
            batch_size: Tamaño de lote
            port: Puerto

        Returns:
            Dict con insertedCount, updatedCount, errorCount, errors

        Example:
            | ${result}= | Load To Database | postgres | localhost | mydb | user | pass | customers | ${records} |
        """
        if not records:
            return LoadResult(0, 0, 0, []).to_dict()

        conn = self._get_connection(db_type, host, database, username, password, port, schema, warehouse)
        cursor = conn.cursor()

        inserted = 0
        errors = []
        columns = list(records[0].keys())

        # Build insert statement
        placeholders = ", ".join(["%s"] * len(columns))
        cols_str = ", ".join(columns)
        insert_sql = f"INSERT INTO {table} ({cols_str}) VALUES ({placeholders})"

        # Insert in batches
        for i in range(0, len(records), batch_size):
            batch = records[i:i + batch_size]
            for record in batch:
                try:
                    values = [record.get(col) for col in columns]
                    cursor.execute(insert_sql, values)
                    inserted += 1
                except Exception as e:
                    errors.append(str(e))

            conn.commit()

        cursor.close()

        return LoadResult(
            inserted_count=inserted,
            updated_count=0,
            error_count=len(errors),
            errors=errors[:10],  # Limit error messages
        ).to_dict()

    # =========================================================================
    # CSV OPERATIONS
    # =========================================================================

    @keyword("Extract From CSV")
    def extract_from_csv(
        self,
        path: str,
        delimiter: str = ",",
        encoding: str = "utf-8",
        header: bool = True,
        columns: Optional[str] = None,
        limit: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Extrae datos de un archivo CSV.

        Args:
            path: Ruta al archivo CSV
            delimiter: Delimitador (default: ,)
            encoding: Codificación del archivo
            header: Si tiene fila de encabezado
            columns: Nombres de columnas personalizados (si header=False)
            limit: Límite de filas

        Returns:
            Dict con records, columns, recordCount

        Example:
            | ${result}= | Extract From CSV | /path/to/file.csv |
        """
        import csv

        records = []
        col_names = []

        with open(path, "r", encoding=encoding) as f:
            if header:
                reader = csv.DictReader(f, delimiter=delimiter)
                for i, row in enumerate(reader):
                    if limit and i >= limit:
                        break
                    records.append(dict(row))
                col_names = reader.fieldnames or []
            else:
                reader = csv.reader(f, delimiter=delimiter)
                if columns:
                    col_names = [c.strip() for c in columns.split(",")]
                for i, row in enumerate(reader):
                    if limit and i >= limit:
                        break
                    if not col_names:
                        col_names = [f"col_{j}" for j in range(len(row))]
                    records.append(dict(zip(col_names, row)))

        return ExtractionResult(
            records=records,
            columns=col_names,
            record_count=len(records),
        ).to_dict()

    @keyword("Load To CSV")
    def load_to_csv(
        self,
        path: str,
        records: List[Dict[str, Any]],
        delimiter: str = ",",
        encoding: str = "utf-8",
        append: bool = False,
        columns: Optional[Any] = None,
    ) -> Dict[str, Any]:
        """
        Guarda datos en un archivo CSV.

        Args:
            path: Ruta al archivo CSV
            records: Lista de diccionarios
            delimiter: Delimitador
            encoding: Codificación
            append: Si agregar al archivo existente
            columns: Constructor opcional de columnas. Formatos soportados:
                - "field1,field2"
                - {"Header1":"field1","Header2":"field2"}
                - ["field1","field2"]
                - [{"header":"Date","field":"fecha","default":""}, ...]

        Returns:
            Dict con insertedCount

        Example:
            | ${result}= | Load To CSV | /path/to/output.csv | ${records} |
        """
        import csv

        if not records:
            return LoadResult(0, 0, 0, []).to_dict()

        output_path = self._resolve_csv_output_path(path)
        mode = "a" if append else "w"
        records_to_write = records

        column_specs = self._parse_csv_columns(columns)
        if column_specs:
            fieldnames = [spec["header"] for spec in column_specs]
            transformed_records: List[Dict[str, Any]] = []
            for record in records:
                row: Dict[str, Any] = {}
                for spec in column_specs:
                    header = spec["header"]
                    default = spec.get("default", "")
                    if "value" in spec:
                        resolved = spec["value"]
                    else:
                        resolved = self._get_nested_value(record, spec.get("field", ""))
                    row[header] = default if resolved is None else resolved
                transformed_records.append(row)
            records_to_write = transformed_records
        else:
            fieldnames = list(records[0].keys())

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open(mode, encoding=encoding, newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter)
            if not append or f.tell() == 0:
                writer.writeheader()
            writer.writerows(records_to_write)

        result = LoadResult(
            inserted_count=len(records_to_write),
            updated_count=0,
            error_count=0,
            errors=[],
        ).to_dict()
        # Surface effective path (especially when caller passed a directory).
        result["path"] = str(output_path)
        return result

    def _resolve_csv_output_path(self, path: str) -> Path:
        """Resolve output path for CSV writes; accept directories as destination."""
        raw = str(path or "").strip()
        if not raw:
            raise ValueError("CSV path cannot be empty")

        candidate = Path(raw).expanduser()
        # If caller passed a directory, write to default file inside it.
        if candidate.exists() and candidate.is_dir():
            return candidate / "output.csv"
        if raw.endswith("/") or raw.endswith("\\"):
            return candidate / "output.csv"
        return candidate

    @keyword("Load To QBO")
    def load_to_qbo(
        self,
        path: str,
        records: List[Dict[str, Any]],
        account_type: str = "CHECKING",
        account_id: str = "000000000",
        bank_id: str = "000000000",
        currency: str = "USD",
        date_field: str = "date",
        amount_field: str = "amount",
        payee_field: str = "name",
        memo_field: str = "memo",
        fitid_field: str = "fitid",
        type_field: str = "type",
        date_format: Optional[str] = None,
        intu_bid: str = "3000",
        encoding: str = "cp1252",
    ) -> Dict[str, Any]:
        """
        Guarda datos en un archivo QBO (OFX) compatible con QuickBooks.

        Args:
            path: Ruta al archivo QBO
            records: Lista de diccionarios con transacciones
            account_type: Tipo de cuenta (CHECKING, SAVINGS, MONEYMRKT, CREDITLINE, CREDITCARD)
            account_id: Identificador de cuenta
            bank_id: Identificador del banco (solo cuentas bancarias)
            currency: Moneda (default: USD)
            date_field: Campo origen para fecha
            amount_field: Campo origen para monto
            payee_field: Campo origen para beneficiario/nombre
            memo_field: Campo origen para memo
            fitid_field: Campo origen para ID único de transacción
            type_field: Campo origen para tipo de transacción (opcional)
            date_format: Formato explícito de fecha (opcional, strptime)
            intu_bid: Intuit Bank ID para QuickBooks
            encoding: Codificación de salida (default: cp1252)

        Returns:
            Dict con insertedCount, errorCount, errors y path
        """
        if records is None:
            records = []
        if not isinstance(records, list):
            raise ValueError("records must be a list of objects")

        output_path = self._resolve_qbo_output_path(path)
        normalized_account_type = self._normalize_qbo_account_type(account_type)
        is_credit_card = normalized_account_type == "CREDITCARD"
        normalized_currency = str(currency or "USD").strip().upper() or "USD"
        normalized_intu_bid = str(intu_bid or "3000").strip() or "3000"

        transactions: List[Dict[str, Any]] = []
        errors: List[str] = []
        for index, record in enumerate(records):
            if not isinstance(record, dict):
                errors.append(f"Record {index} is not an object")
                continue
            try:
                tx = self._normalize_qbo_transaction(
                    record=record,
                    index=index,
                    date_field=date_field,
                    amount_field=amount_field,
                    payee_field=payee_field,
                    memo_field=memo_field,
                    fitid_field=fitid_field,
                    type_field=type_field,
                    date_format=date_format,
                )
                transactions.append(tx)
            except Exception as exc:
                errors.append(f"Record {index}: {exc}")

        if records and not transactions:
            preview = "; ".join(errors[:3]) if errors else "No valid transactions."
            raise ValueError(
                "No valid QBO transactions were produced from input records. "
                f"Sample errors: {preview}"
            )

        qbo_content = self._build_qbo_document(
            transactions=transactions,
            account_type=normalized_account_type,
            account_id=str(account_id or "000000000").strip() or "000000000",
            bank_id=str(bank_id or "000000000").strip() or "000000000",
            currency=normalized_currency,
            intu_bid=normalized_intu_bid,
            is_credit_card=is_credit_card,
        )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open("w", encoding=encoding, newline="\n") as f:
            f.write(qbo_content)

        result = LoadResult(
            inserted_count=len(transactions),
            updated_count=0,
            error_count=len(errors),
            errors=errors,
        ).to_dict()
        result["path"] = str(output_path)
        return result

    def _resolve_qbo_output_path(self, path: str) -> Path:
        """Resolve output path for QBO writes; accept directories as destination."""
        raw = str(path or "").strip()
        if not raw:
            raise ValueError("QBO path cannot be empty")

        candidate = Path(raw).expanduser()
        if candidate.exists() and candidate.is_dir():
            return candidate / "output.qbo"
        if raw.endswith("/") or raw.endswith("\\"):
            return candidate / "output.qbo"
        return candidate

    def _normalize_qbo_account_type(self, account_type: str) -> str:
        normalized = str(account_type or "").strip().upper()
        if normalized in {"CREDITCARD", "CREDIT_CARD", "CCARD"}:
            return "CREDITCARD"
        if normalized in {"CHECKING", "SAVINGS", "MONEYMRKT", "CREDITLINE"}:
            return normalized
        return "CHECKING"

    def _normalize_qbo_transaction(
        self,
        record: Dict[str, Any],
        index: int,
        date_field: str,
        amount_field: str,
        payee_field: str,
        memo_field: str,
        fitid_field: str,
        type_field: str,
        date_format: Optional[str],
    ) -> Dict[str, Any]:
        raw_date = self._get_qbo_field_value(
            record,
            primary=date_field,
            aliases=["txn_date", "transaction_date", "posted_date", "fecha"],
        )
        if raw_date is None:
            raise ValueError(f"missing date field '{date_field}'")
        posted_date = self._coerce_qbo_date(raw_date, date_format)

        raw_amount = self._get_qbo_field_value(
            record,
            primary=amount_field,
            aliases=["txn_amount", "transaction_amount", "monto", "importe"],
        )
        if raw_amount is None:
            raise ValueError(f"missing amount field '{amount_field}'")
        amount = self._coerce_qbo_amount(raw_amount)

        raw_payee = self._get_qbo_field_value(
            record,
            primary=payee_field,
            aliases=["payee", "description", "merchant", "beneficiary", "nombre"],
        )
        raw_memo = self._get_qbo_field_value(
            record,
            primary=memo_field,
            aliases=["description", "details", "note", "notes", "memo_text"],
        )
        raw_type = (
            self._get_qbo_field_value(
                record,
                primary=type_field,
                aliases=["transaction_type", "txn_type", "trn_type"],
            )
            if type_field
            else None
        )
        raw_fitid = (
            self._get_qbo_field_value(
                record,
                primary=fitid_field,
                aliases=["id", "transaction_id", "txn_id", "reference", "ref"],
            )
            if fitid_field
            else None
        )

        payee = self._sanitize_qbo_text(raw_payee)
        memo = self._sanitize_qbo_text(raw_memo)
        trn_type = self._coerce_qbo_trn_type(raw_type, amount)
        fitid = self._coerce_qbo_fitid(
            raw_fitid=raw_fitid,
            index=index,
            posted_date=posted_date,
            amount=amount,
            payee=payee,
            memo=memo,
        )

        return {
            "trn_type": trn_type,
            "dt_posted": posted_date.strftime("%Y%m%d000000"),
            "amount": self._format_qbo_amount(amount),
            "fitid": fitid,
            "name": payee,
            "memo": memo,
            "date_obj": posted_date,
            "amount_obj": amount,
        }

    def _get_qbo_field_value(
        self,
        record: Dict[str, Any],
        primary: str,
        aliases: Optional[List[str]] = None,
    ) -> Any:
        """Resolve QBO field with tolerant matching:
        1) explicit field (exact, then case-insensitive)
        2) common aliases (exact, then case-insensitive)
        """
        candidates: List[str] = []
        if primary:
            candidates.append(str(primary).strip())

        for alias in aliases or []:
            alias_text = str(alias).strip()
            if alias_text:
                candidates.append(alias_text)

        seen: set[str] = set()
        for candidate in candidates:
            key = candidate.lower()
            if not candidate or key in seen:
                continue
            seen.add(key)

            value = self._get_nested_value(record, candidate, case_insensitive=False)
            if value is not None:
                return value

            value = self._get_nested_value(record, candidate, case_insensitive=True)
            if value is not None:
                return value

        return None

    def _coerce_qbo_date(self, value: Any, date_format: Optional[str]) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            return datetime.fromtimestamp(float(value), tz=timezone.utc).date()

        text = str(value or "").strip()
        if not text:
            raise ValueError("empty date")

        if date_format:
            return datetime.strptime(text, date_format).date()

        candidate = text.replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(candidate).date()
        except ValueError:
            pass

        known_formats = [
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%Y.%m.%d",
            "%m/%d/%Y",
            "%m-%d-%Y",
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%m/%d/%y",
            "%m-%d-%y",
            "%d/%m/%y",
            "%d-%m-%y",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
            "%m/%d/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
            "%m/%d/%y %H:%M:%S",
            "%d/%m/%y %H:%M:%S",
        ]
        for fmt in known_formats:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue

        raise ValueError(f"unsupported date format '{text}'")

    def _coerce_qbo_amount(self, value: Any) -> Decimal:
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))

        text = str(value or "").strip()
        if not text:
            raise ValueError("empty amount")

        negative_parentheses = text.startswith("(") and text.endswith(")")
        cleaned = text.replace("(", "").replace(")", "")
        cleaned = cleaned.replace("$", "").replace(",", "").strip()
        if cleaned.endswith("-"):
            cleaned = f"-{cleaned[:-1]}"

        try:
            amount = Decimal(cleaned)
        except InvalidOperation as exc:
            raise ValueError(f"invalid amount '{value}'") from exc

        if negative_parentheses and amount > 0:
            amount = -amount
        return amount

    def _coerce_qbo_trn_type(self, raw_type: Any, amount: Decimal) -> str:
        normalized = str(raw_type or "").strip().upper()
        if normalized:
            debit_aliases = {
                "DEBIT",
                "PAYMENT",
                "WITHDRAWAL",
                "CHECK",
                "POS",
                "ATM",
                "FEE",
                "PURCHASE",
            }
            credit_aliases = {"CREDIT", "DEPOSIT", "REFUND", "DIRECTDEP", "DIV", "INT"}
            if normalized in debit_aliases:
                return "DEBIT"
            if normalized in credit_aliases:
                return "CREDIT"
        return "DEBIT" if amount < 0 else "CREDIT"

    def _coerce_qbo_fitid(
        self,
        raw_fitid: Any,
        index: int,
        posted_date: date,
        amount: Decimal,
        payee: str,
        memo: str,
    ) -> str:
        text = self._sanitize_qbo_text(raw_fitid, limit=64)
        if text:
            return text
        seed = f"{index}|{posted_date.isoformat()}|{amount}|{payee}|{memo}"
        return hashlib.sha1(seed.encode("utf-8")).hexdigest()[:32]

    def _format_qbo_amount(self, amount: Decimal) -> str:
        return format(amount.quantize(Decimal("0.01")), "f")

    def _sanitize_qbo_text(self, value: Any, limit: int = 128) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        compact = " ".join(text.splitlines()).strip()
        return compact[:limit]

    def _build_qbo_document(
        self,
        transactions: List[Dict[str, Any]],
        account_type: str,
        account_id: str,
        bank_id: str,
        currency: str,
        intu_bid: str,
        is_credit_card: bool,
    ) -> str:
        now_utc = datetime.now(timezone.utc)
        dt_server = now_utc.strftime("%Y%m%d%H%M%S")

        if transactions:
            dt_start = min(t["date_obj"] for t in transactions).strftime("%Y%m%d000000")
            dt_end = max(t["date_obj"] for t in transactions).strftime("%Y%m%d000000")
            ledger_amount = sum((t["amount_obj"] for t in transactions), Decimal("0.00"))
        else:
            fallback = now_utc.strftime("%Y%m%d000000")
            dt_start = fallback
            dt_end = fallback
            ledger_amount = Decimal("0.00")

        tx_lines: List[str] = []
        for tx in transactions:
            tx_lines.extend(
                [
                    "<STMTTRN>",
                    f"<TRNTYPE>{tx['trn_type']}",
                    f"<DTPOSTED>{tx['dt_posted']}",
                    f"<TRNAMT>{tx['amount']}",
                    f"<FITID>{tx['fitid']}",
                    f"<NAME>{tx['name']}",
                    f"<MEMO>{tx['memo']}",
                    "</STMTTRN>",
                ]
            )

        uid_seed = f"{dt_server}|{account_id}|{len(transactions)}"
        new_uid = hashlib.md5(uid_seed.encode("utf-8")).hexdigest().upper()

        lines: List[str] = [
            "OFXHEADER:100",
            "DATA:OFXSGML",
            "VERSION:102",
            "SECURITY:NONE",
            "ENCODING:USASCII",
            "CHARSET:1252",
            "COMPRESSION:NONE",
            "OLDFILEUID:NONE",
            f"NEWFILEUID:{new_uid}",
            "",
            "<OFX>",
            "<SIGNONMSGSRSV1>",
            "<SONRS>",
            "<STATUS>",
            "<CODE>0",
            "<SEVERITY>INFO",
            "</STATUS>",
            f"<DTSERVER>{dt_server}",
            "<LANGUAGE>ENG",
            "</SONRS>",
            "</SIGNONMSGSRSV1>",
        ]

        if is_credit_card:
            lines.extend(
                [
                    "<CREDITCARDMSGSRSV1>",
                    "<CCSTMTTRNRS>",
                    "<TRNUID>0",
                    "<STATUS>",
                    "<CODE>0",
                    "<SEVERITY>INFO",
                    "</STATUS>",
                    "<CCSTMTRS>",
                    f"<CURDEF>{currency}",
                    "<CCACCTFROM>",
                    f"<ACCTID>{account_id}",
                    "</CCACCTFROM>",
                    f"<INTU.BID>{intu_bid}",
                    "<BANKTRANLIST>",
                    f"<DTSTART>{dt_start}",
                    f"<DTEND>{dt_end}",
                ]
            )
        else:
            lines.extend(
                [
                    "<BANKMSGSRSV1>",
                    "<STMTTRNRS>",
                    "<TRNUID>0",
                    "<STATUS>",
                    "<CODE>0",
                    "<SEVERITY>INFO",
                    "</STATUS>",
                    "<STMTRS>",
                    f"<CURDEF>{currency}",
                    "<BANKACCTFROM>",
                    f"<BANKID>{bank_id}",
                    f"<ACCTID>{account_id}",
                    f"<ACCTTYPE>{account_type}",
                    "</BANKACCTFROM>",
                    f"<INTU.BID>{intu_bid}",
                    "<BANKTRANLIST>",
                    f"<DTSTART>{dt_start}",
                    f"<DTEND>{dt_end}",
                ]
            )

        lines.extend(tx_lines)
        lines.extend(
            [
                "</BANKTRANLIST>",
                "<LEDGERBAL>",
                f"<BALAMT>{self._format_qbo_amount(ledger_amount)}",
                f"<DTASOF>{dt_end}",
                "</LEDGERBAL>",
            ]
        )

        if is_credit_card:
            lines.extend(
                [
                    "</CCSTMTRS>",
                    "</CCSTMTTRNRS>",
                    "</CREDITCARDMSGSRSV1>",
                ]
            )
        else:
            lines.extend(
                [
                    "</STMTRS>",
                    "</STMTTRNRS>",
                    "</BANKMSGSRSV1>",
                ]
            )

        lines.append("</OFX>")
        return "\n".join(lines) + "\n"

    def _parse_csv_columns(self, columns: Optional[Any]) -> List[Dict[str, Any]]:
        """Parse CSV constructor config into normalized column specs."""
        if columns is None:
            return []

        parsed: Any = columns
        if isinstance(columns, str):
            raw = columns.strip()
            if not raw:
                return []
            if raw[0] in ("[", "{"):
                parsed = json.loads(raw)
            else:
                parsed = [part.strip() for part in raw.split(",") if part.strip()]

        specs: List[Dict[str, Any]] = []
        if isinstance(parsed, dict):
            for header, field in parsed.items():
                header_name = str(header).strip()
                field_name = str(field).strip() if field is not None else ""
                if not header_name:
                    continue
                specs.append({"header": header_name, "field": field_name or header_name})
            return specs

        if not isinstance(parsed, list):
            raise ValueError("CSV columns must be list/dict/string")

        for entry in parsed:
            if isinstance(entry, str):
                field_name = entry.strip()
                if not field_name:
                    continue
                specs.append({"header": field_name, "field": field_name})
                continue

            if not isinstance(entry, dict):
                raise ValueError("CSV column entries must be strings or objects")

            header = entry.get("header") or entry.get("name") or entry.get("column")
            field = entry.get("field") or entry.get("source") or entry.get("path") or entry.get("key")
            default = entry.get("default", "")
            value = entry.get("value")

            header_name = str(header).strip() if header is not None else ""
            field_name = str(field).strip() if field is not None else ""
            if not header_name:
                header_name = field_name
            if not header_name:
                raise ValueError("CSV column object requires header/name/column or field/source/path/key")

            spec: Dict[str, Any] = {"header": header_name, "default": default}
            if value is not None:
                spec["value"] = value
            else:
                spec["field"] = field_name or header_name
            specs.append(spec)

        return specs

    def _get_nested_value(
        self,
        record: Dict[str, Any],
        path: str,
        case_insensitive: bool = False,
    ) -> Any:
        """Get nested value from dict/list with dot notation and array indexes."""
        if not path:
            return None

        current: Any = record
        for part in str(path).split("."):
            key = part.strip()
            if key == "":
                return None

            if isinstance(current, dict):
                if key in current:
                    current = current[key]
                    continue

                if not case_insensitive:
                    return None

                lowered = key.lower()
                match_key = next(
                    (existing for existing in current.keys() if str(existing).lower() == lowered),
                    None,
                )
                if match_key is None:
                    return None
                current = current[match_key]
                continue

            if isinstance(current, list):
                if not key.isdigit():
                    return None
                idx = int(key)
                if idx < 0 or idx >= len(current):
                    return None
                current = current[idx]
                continue

            return None

        return current

    # =========================================================================
    # EXCEL OPERATIONS (delegates to ExcelLibrary for file ops)
    # =========================================================================

    @keyword("Extract From Excel")
    def extract_from_excel(
        self,
        path: str,
        sheet: Optional[str] = None,
        header: bool = True,
        columns: Optional[str] = None,
        limit: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Extrae datos de un archivo Excel.

        Args:
            path: Ruta al archivo Excel
            sheet: Nombre de la hoja
            header: Si tiene fila de encabezado
            columns: Nombres de columnas personalizados
            limit: Límite de filas

        Returns:
            Dict con records, columns, recordCount

        Example:
            | ${result}= | Extract From Excel | /path/to/file.xlsx | Sheet1 |
        """
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active

        records = []
        col_names = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                if header:
                    col_names = [str(c) if c else f"col_{j}" for j, c in enumerate(row)]
                    continue
                elif columns:
                    col_names = [c.strip() for c in columns.split(",")]
                else:
                    col_names = [f"col_{j}" for j in range(len(row))]

            if limit and len(records) >= limit:
                break

            records.append(dict(zip(col_names, row)))

        wb.close()

        return ExtractionResult(
            records=records,
            columns=col_names,
            record_count=len(records),
        ).to_dict()

    @keyword("Load To Excel")
    def load_to_excel(
        self,
        path: str,
        records: List[Dict[str, Any]],
        sheet: str = "Sheet1",
    ) -> Dict[str, Any]:
        """
        Guarda datos en un archivo Excel.

        Args:
            path: Ruta al archivo Excel
            records: Lista de diccionarios
            sheet: Nombre de la hoja

        Returns:
            Dict con insertedCount

        Example:
            | ${result}= | Load To Excel | /path/to/output.xlsx | ${records} |
        """
        import openpyxl

        if not records:
            return LoadResult(0, 0, 0, []).to_dict()

        output_path = self._resolve_excel_output_path(path)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet

        # Write headers
        fieldnames = list(records[0].keys())
        for col, name in enumerate(fieldnames, start=1):
            ws.cell(row=1, column=col, value=name)

        # Write data
        for row_idx, record in enumerate(records, start=2):
            for col_idx, key in enumerate(fieldnames, start=1):
                ws.cell(row=row_idx, column=col_idx, value=record.get(key))

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

        result = LoadResult(
            inserted_count=len(records),
            updated_count=0,
            error_count=0,
            errors=[],
        ).to_dict()
        result["path"] = str(output_path)
        return result

    def _resolve_excel_output_path(self, path: str) -> Path:
        """Resolve output path for Excel writes; accept directories as destination."""
        raw = str(path or "").strip()
        if not raw:
            raise ValueError("Excel path cannot be empty")

        candidate = Path(raw).expanduser()
        if candidate.exists() and candidate.is_dir():
            return candidate / "output.xlsx"
        if raw.endswith("/") or raw.endswith("\\"):
            return candidate / "output.xlsx"
        return candidate

    # =========================================================================
    # S3 OPERATIONS
    # =========================================================================

    @keyword("Extract From S3")
    def extract_from_s3(
        self,
        bucket: str,
        key: str,
        aws_access_key: str,
        aws_secret_key: str,
        region: str = "us-east-1",
        file_type: str = "csv",
    ) -> Dict[str, Any]:
        """
        Extrae datos de un archivo en S3.

        Args:
            bucket: Nombre del bucket
            key: Path del archivo en S3
            aws_access_key: AWS Access Key ID
            aws_secret_key: AWS Secret Access Key
            region: Región AWS
            file_type: Tipo de archivo (csv, json)

        Returns:
            Dict con records, columns, recordCount

        Example:
            | ${result}= | Extract From S3 | my-bucket | data/file.csv | ${KEY} | ${SECRET} |
        """
        import boto3
        import tempfile
        import os

        s3 = boto3.client(
            "s3",
            aws_access_key_id=aws_access_key,
            aws_secret_access_key=aws_secret_key,
            region_name=region,
        )

        # Download to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_type}") as tmp:
            s3.download_file(bucket, key, tmp.name)
            tmp_path = tmp.name

        try:
            if file_type == "csv":
                result = self.extract_from_csv(tmp_path)
            elif file_type == "json":
                with open(tmp_path, "r") as f:
                    data = json.load(f)
                if isinstance(data, list):
                    records = data
                else:
                    records = [data]
                columns = list(records[0].keys()) if records else []
                result = ExtractionResult(
                    records=records,
                    columns=columns,
                    record_count=len(records),
                ).to_dict()
            else:
                raise ValueError(f"Tipo de archivo no soportado: {file_type}")
        finally:
            os.unlink(tmp_path)

        return result

    @keyword("Load To S3")
    def load_to_s3(
        self,
        bucket: str,
        key: str,
        records: List[Dict[str, Any]],
        aws_access_key: str,
        aws_secret_key: str,
        region: str = "us-east-1",
        file_type: str = "csv",
    ) -> Dict[str, Any]:
        """
        Sube datos a S3.

        Args:
            bucket: Nombre del bucket
            key: Path destino en S3
            records: Lista de diccionarios
            aws_access_key: AWS Access Key ID
            aws_secret_key: AWS Secret Access Key
            region: Región AWS
            file_type: Tipo de archivo (csv, json)

        Returns:
            Dict con insertedCount

        Example:
            | ${result}= | Load To S3 | my-bucket | output/data.csv | ${records} | ${KEY} | ${SECRET} |
        """
        import boto3
        import tempfile
        import os

        s3 = boto3.client(
            "s3",
            aws_access_key_id=aws_access_key,
            aws_secret_access_key=aws_secret_key,
            region_name=region,
        )

        # Write to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_type}") as tmp:
            tmp_path = tmp.name

        try:
            if file_type == "csv":
                self.load_to_csv(tmp_path, records)
            elif file_type == "json":
                with open(tmp_path, "w") as f:
                    json.dump(records, f)
            else:
                raise ValueError(f"Tipo de archivo no soportado: {file_type}")

            s3.upload_file(tmp_path, bucket, key)
        finally:
            os.unlink(tmp_path)

        return LoadResult(
            inserted_count=len(records),
            updated_count=0,
            error_count=0,
            errors=[],
        ).to_dict()

    # =========================================================================
    # SFTP OPERATIONS
    # =========================================================================

    @keyword("Extract From SFTP")
    def extract_from_sftp(
        self,
        host: str,
        path: str,
        username: str,
        password: Optional[str] = None,
        private_key: Optional[str] = None,
        port: int = 22,
        file_type: str = "csv",
    ) -> Dict[str, Any]:
        """
        Extrae datos de un archivo en SFTP.

        Args:
            host: Host SFTP
            path: Path del archivo
            username: Usuario
            password: Contraseña (o usar private_key)
            private_key: Path a llave privada
            port: Puerto SFTP
            file_type: Tipo de archivo (csv, json)

        Returns:
            Dict con records, columns, recordCount

        Example:
            | ${result}= | Extract From SFTP | sftp.example.com | /data/file.csv | user | pass |
        """
        import paramiko
        import tempfile
        import os

        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        if private_key:
            pkey = paramiko.RSAKey.from_private_key_file(private_key)
            ssh.connect(host, port=port, username=username, pkey=pkey)
        else:
            ssh.connect(host, port=port, username=username, password=password)

        sftp = ssh.open_sftp()

        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_type}") as tmp:
            sftp.get(path, tmp.name)
            tmp_path = tmp.name

        sftp.close()
        ssh.close()

        try:
            if file_type == "csv":
                result = self.extract_from_csv(tmp_path)
            elif file_type == "json":
                with open(tmp_path, "r") as f:
                    data = json.load(f)
                records = data if isinstance(data, list) else [data]
                columns = list(records[0].keys()) if records else []
                result = ExtractionResult(
                    records=records,
                    columns=columns,
                    record_count=len(records),
                ).to_dict()
            else:
                raise ValueError(f"Tipo de archivo no soportado: {file_type}")
        finally:
            os.unlink(tmp_path)

        return result

    @keyword("Load To SFTP")
    def load_to_sftp(
        self,
        host: str,
        path: str,
        records: List[Dict[str, Any]],
        username: str,
        password: Optional[str] = None,
        private_key: Optional[str] = None,
        port: int = 22,
        file_type: str = "csv",
    ) -> Dict[str, Any]:
        """
        Sube datos a SFTP.

        Args:
            host: Host SFTP
            path: Path destino
            records: Lista de diccionarios
            username: Usuario
            password: Contraseña
            private_key: Path a llave privada
            port: Puerto
            file_type: Tipo de archivo

        Returns:
            Dict con insertedCount

        Example:
            | ${result}= | Load To SFTP | sftp.example.com | /output/data.csv | ${records} | user | pass |
        """
        import paramiko
        import tempfile
        import os

        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        if private_key:
            pkey = paramiko.RSAKey.from_private_key_file(private_key)
            ssh.connect(host, port=port, username=username, pkey=pkey)
        else:
            ssh.connect(host, port=port, username=username, password=password)

        sftp = ssh.open_sftp()

        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_type}") as tmp:
            tmp_path = tmp.name

        try:
            if file_type == "csv":
                self.load_to_csv(tmp_path, records)
            elif file_type == "json":
                with open(tmp_path, "w") as f:
                    json.dump(records, f)
            else:
                raise ValueError(f"Tipo de archivo no soportado: {file_type}")

            sftp.put(tmp_path, path)
        finally:
            os.unlink(tmp_path)

        sftp.close()
        ssh.close()

        return LoadResult(
            inserted_count=len(records),
            updated_count=0,
            error_count=0,
            errors=[],
        ).to_dict()

    # =========================================================================
    # REST API OPERATIONS
    # =========================================================================

    @keyword("Extract From REST API")
    def extract_from_rest_api(
        self,
        url: str,
        method: str = "GET",
        headers: Optional[str] = None,
        body: Optional[str] = None,
        auth_type: Optional[str] = None,
        auth_value: Optional[str] = None,
        pagination_type: Optional[str] = None,
        pagination_param: Optional[str] = None,
        data_path: Optional[str] = None,
        limit: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Extrae datos de una REST API con soporte para paginación.

        Args:
            url: URL del endpoint
            method: Método HTTP (GET, POST)
            headers: Headers como JSON string
            body: Body como JSON string (para POST)
            auth_type: Tipo de auth (bearer, basic, api_key)
            auth_value: Valor de auth (token, user:pass, key)
            pagination_type: Tipo de paginación (offset, cursor, link)
            pagination_param: Parámetro de paginación
            data_path: JSONPath al array de datos (ej: "$.data", "$.results")
            limit: Límite de records

        Returns:
            Dict con records, columns, recordCount

        Example:
            | ${result}= | Extract From REST API | https://api.example.com/users | headers={"Accept": "application/json"} |
        """
        import requests

        req_headers = json.loads(headers) if headers else {}

        # Add authentication
        if auth_type == "bearer":
            req_headers["Authorization"] = f"Bearer {auth_value}"
        elif auth_type == "api_key":
            req_headers["X-API-Key"] = auth_value

        all_records = []
        page = 1
        offset = 0
        next_cursor = None

        while True:
            # Build URL with pagination
            paginated_url = url
            if pagination_type == "offset":
                sep = "&" if "?" in url else "?"
                paginated_url = f"{url}{sep}{pagination_param}={offset}"
            elif pagination_type == "page":
                sep = "&" if "?" in url else "?"
                paginated_url = f"{url}{sep}{pagination_param}={page}"
            elif pagination_type == "cursor" and next_cursor:
                sep = "&" if "?" in url else "?"
                paginated_url = f"{url}{sep}{pagination_param}={next_cursor}"

            # Make request
            if method.upper() == "GET":
                response = requests.get(paginated_url, headers=req_headers)
            else:
                req_body = json.loads(body) if body else None
                response = requests.post(paginated_url, headers=req_headers, json=req_body)

            response.raise_for_status()
            data = response.json()

            # Extract records from response
            if data_path:
                # Simple JSONPath extraction
                path_parts = data_path.replace("$.", "").split(".")
                records = data
                for part in path_parts:
                    if part and records:
                        records = records.get(part, [])
            else:
                records = data if isinstance(data, list) else [data]

            if not records:
                break

            all_records.extend(records)

            # Check limit
            if limit and len(all_records) >= limit:
                all_records = all_records[:limit]
                break

            # Handle pagination
            if pagination_type == "offset":
                offset += len(records)
                if len(records) < 100:  # Assume page size of 100
                    break
            elif pagination_type == "page":
                page += 1
                if len(records) < 100:
                    break
            elif pagination_type == "cursor":
                next_cursor = data.get("next_cursor") or data.get("cursor")
                if not next_cursor:
                    break
            else:
                break  # No pagination

        columns = list(all_records[0].keys()) if all_records else []

        return ExtractionResult(
            records=all_records,
            columns=columns,
            record_count=len(all_records),
        ).to_dict()

    # =========================================================================
    # SALESFORCE OPERATIONS
    # =========================================================================

    @keyword("Extract From Salesforce")
    def extract_from_salesforce(
        self,
        username: str,
        password: str,
        security_token: str,
        query: str,
        domain: str = "login",
    ) -> Dict[str, Any]:
        """
        Extrae datos de Salesforce usando SOQL.

        Args:
            username: Usuario Salesforce
            password: Contraseña
            security_token: Security Token
            query: Query SOQL
            domain: Dominio (login, test)

        Returns:
            Dict con records, columns, recordCount

        Example:
            | ${result}= | Extract From Salesforce | user@company.com | pass | token | SELECT Id, Name FROM Account |
        """
        from simple_salesforce import Salesforce

        sf = Salesforce(
            username=username,
            password=password,
            security_token=security_token,
            domain=domain,
        )

        result = sf.query_all(query)
        records = result.get("records", [])

        # Remove Salesforce metadata from records
        clean_records = []
        for record in records:
            clean = {k: v for k, v in record.items() if not k.startswith("attributes")}
            clean_records.append(clean)

        columns = list(clean_records[0].keys()) if clean_records else []

        return ExtractionResult(
            records=clean_records,
            columns=columns,
            record_count=len(clean_records),
        ).to_dict()

    # =========================================================================
    # BIGQUERY OPERATIONS (Target only)
    # =========================================================================

    @keyword("Load To BigQuery")
    def load_to_bigquery(
        self,
        project: str,
        dataset: str,
        table: str,
        records: List[Dict[str, Any]],
        credentials_json: str,
        mode: str = "append",
    ) -> Dict[str, Any]:
        """
        Carga datos en Google BigQuery.

        Args:
            project: ID del proyecto GCP
            dataset: Dataset de BigQuery
            table: Tabla destino
            records: Lista de diccionarios
            credentials_json: Path al archivo de credenciales JSON
            mode: Modo de escritura (append, truncate)

        Returns:
            Dict con insertedCount

        Example:
            | ${result}= | Load To BigQuery | my-project | my_dataset | my_table | ${records} | /path/to/creds.json |
        """
        from google.cloud import bigquery
        from google.oauth2 import service_account

        credentials = service_account.Credentials.from_service_account_file(
            credentials_json
        )

        client = bigquery.Client(credentials=credentials, project=project)

        table_id = f"{project}.{dataset}.{table}"

        write_disposition = (
            bigquery.WriteDisposition.WRITE_APPEND
            if mode == "append"
            else bigquery.WriteDisposition.WRITE_TRUNCATE
        )

        job_config = bigquery.LoadJobConfig(
            write_disposition=write_disposition,
            source_format=bigquery.SourceFormat.NEWLINE_DELIMITED_JSON,
        )

        job = client.load_table_from_json(records, table_id, job_config=job_config)
        job.result()  # Wait for completion

        return LoadResult(
            inserted_count=len(records),
            updated_count=0,
            error_count=0,
            errors=[],
        ).to_dict()

    # =========================================================================
    # CLEANUP
    # =========================================================================

    @keyword("Close All Data Connections")
    def close_all_connections(self):
        """
        Cierra todas las conexiones de base de datos abiertas.

        Example:
            | Close All Data Connections |
        """
        for key, conn in self._connections.items():
            try:
                conn.close()
            except Exception:
                pass
        self._connections.clear()
